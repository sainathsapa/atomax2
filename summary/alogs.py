# from ufl import replace
import datetime
import yake
from sentence_splitter import SentenceSplitter, split_text_into_sentences
from transformers import PegasusForConditionalGeneration, PegasusTokenizer, AutoTokenizer, AutoModelForSeq2SeqLM
from googletrans import Translator
import language_tool_python
import PyPDF2
import docx
import regex as re
import string
import PyPDF2
import spacy
import torch
from string import punctuation
# from numba import jit, cuda
from deepmultilingualpunctuation import PunctuationModel
from newspaper import Article
from spacy.lang.en.stop_words import STOP_WORDS
from transformers import (
    PegasusForConditionalGeneration, PegasusTokenizer, pipeline, BartForConditionalGeneration, BartTokenizer)
from heapq import nlargest
# from GPUtil import showUtilization as gpu_usage
import speech_recognition as sr
from os import path
import os
from bs4 import BeautifulSoup
import nltk
from nltk.tokenize import sent_tokenize
from tabula import read_pdf
from tabulate import tabulate
import pandas as pd
import io
import camelot
from diffusers import StableDiffusionPipeline
import datetime
import os
import openai
from nltk.sentiment.vader import SentimentIntensityAnalyzer
import openpyxl
import tweepy
import string
import random
import whisper

import re
from datetime import datetime, timedelta

from pydub import AudioSegment
from pydub.playback import play
import pyaudio

import subprocess

import torch
import pyannote.audio
from pyannote.audio.pipelines.speaker_verification import PretrainedSpeakerEmbedding
from pyannote.audio import Audio
from pyannote.core import Segment

import wave
import contextlib

from sklearn.cluster import AgglomerativeClustering
import numpy as np

fileDIR = "files/"

# ----------------------------------------------------------------
# Translators
# ----------------------------------------------------------------


def translate(summaryText, lang_code):
    translator = Translator()

    Transtxt = translator.translate(summaryText, dest=lang_code)
    return Transtxt


# ----------------------------------------------------------------
#  Convert
# ----------------------------------------------------------------


def filePDF_extract(filePath):

    pdffileobj = open(filePath, 'rb')
    pdfreader = PyPDF2.PdfFileReader(pdffileobj)
    x = pdfreader.numPages
    content = ""
    RetText = ""
    for i in range(x):
        pageobj = pdfreader.getPage(i)
        content = pageobj.extractText()
        RetText = RetText + content
    return RetText


def fileDOCX_extract(filePath):
    doc = docx.Document(filePath)
    fullText = []
    for para in doc.paragraphs:
        fullText.append(para.text)
    return '\n'.join(fullText)


def fileTXT_extract(filepath):
    TextFile = open(filepath, "r")
    textRead = TextFile.read()
    TextFile.close()
    return textRead


def URL_extract(url):

    link_paper = (url)
    article = Article(link_paper)
    article.download()
    article.parse()
    input_text = article.text
    input_text = ' '.join(input_text.split())
    return input_text


def setGrammer(inputText):
    model = PunctuationModel()

    inputText = inputText.replace("<n>", "")

    inputText = model.restore_punctuation(inputText)

    my_tool = language_tool_python.LanguageTool('en-US')

    inputText = my_tool.correct(inputText)

    return inputText


def Clean_Text(textforCleanup):
    text = re.sub("https?:\/\/.*[\r\n]*", "", textforCleanup)
    text = re.sub("#", "", text)
    punct = set(string.punctuation)
    text = "".join([ch for ch in text if ch not in punct])
    text = text.encode(encoding="ascii", errors="ignore")
    text = text.decode()
    clean_text = " ".join([word for word in text.split()])
    clean_text = str(clean_text.lower())
    return clean_text


def fromMic(lang_input):
    whole_text = ''
    r = sr.Recognizer()
    breakwordList = ['stop', 'Stop', 'STOP', 'brake', 'Brake', 'BRAKE']

    while (1):

        try:

            with sr.Microphone(device_index=1) as source:
                r.adjust_for_ambient_noise(source)
                audio = r.listen(source)
                MyText = r.recognize_google(audio, language=lang_input)
                MyText = MyText.lower()
                whole_text += MyText+' '
                print(MyText)
                transText = translate(MyText, 'en').text
                print('---------------')
                # print(transText)
                if transText in breakwordList:
                    return whole_text

        except sr.RequestError as e:
            print("Could not request results; {0}".format(e))

        except sr.UnknownValueError:
            print("unknown error occurred")


# ----------------------------------------------------------------
# BASIC ALGO
# ----------------------------------------------------------------
device = "cpu"


def PegasusModel(spacy_summary):
    model_name = "google/pegasus-large"

    # Load pretrained tokenizer
    # device = "cuda" if torch.cuda.is_available() else "cpu"

    pegasus_tokenizer = PegasusTokenizer.from_pretrained(model_name)

    pegasus_model = PegasusForConditionalGeneration.from_pretrained(
        model_name)

    # Create tokens

    tokens = pegasus_tokenizer(
        spacy_summary, truncation=True, padding="max_length", return_tensors="pt")

    encoded_summary = pegasus_model.generate(**tokens)

    # Decode summarized text

    decoded_summary = pegasus_tokenizer.decode(

        encoded_summary[0],

        skip_special_tokens=True

    )

    # print(decoded_summary)

    summarizer = pipeline(

        "summarization",

        model=model_name,

        tokenizer=pegasus_tokenizer,

        framework="pt", truncation=True

    )

    summary = summarizer(spacy_summary, min_length=30, max_length=150)

    summary = summary[0]["summary_text"]

    summary = summary.replace("<n>", "")

    summary = summary.replace("  ", "")

    print(summary)
    return summary


def parse_text_to_new_article(inputText):
    # return get_response_from_text(" ".join([my_paraphrase(sent) for sent in sent_tokenize(inputText)]))
    openai.api_key = "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"

    response = openai.Completion.create(
        model="text-davinci-003",
        prompt=inputText,
        temperature=0,
        max_tokens=1024,
        top_p=1.0,
        frequency_penalty=0.0,
        presence_penalty=0.0
    )
    return response.get('choices')[0].get('text')


def Spacy(input_text):
    clean_text = str(input_text.lower())
    nlp = spacy.load('en_core_web_sm')
    doc = nlp(clean_text)
    per = 1.0
    tokens = [token.text for token in doc]

    word_frequencies = {}

    for word in doc:

        if word.text.lower() not in list(STOP_WORDS):

            if word.text.lower() not in punctuation:

                if word.text not in word_frequencies.keys():

                    word_frequencies[word.text] = 1

                else:

                    word_frequencies[word.text] += 1

    max_frequency = max(word_frequencies.values())

    for word in word_frequencies.keys():
        word_frequencies[word] = word_frequencies[word] / max_frequency

    sentence_tokens = [sent for sent in doc.sents]
    sentence_scores = {}

    for sent in sentence_tokens:

        for word in sent:

            if word.text.lower() in word_frequencies.keys():

                if sent not in sentence_scores.keys():

                    sentence_scores[sent] = word_frequencies[word.text.lower()]

                else:

                    sentence_scores[sent] += word_frequencies[word.text.lower()]
    select_length = int(len(sentence_tokens) * per)
    summary = nlargest(select_length, sentence_scores, key=sentence_scores.get)
    # print(summary)
    final_summary = [word.text for word in summary]
    spacy_summary = ' '.join(final_summary)
    return spacy_summary


def Brat(input_text, min_len, mx_len):
    model = "Bart"
    bart_model = BartForConditionalGeneration.from_pretrained(
        "facebook/bart-large-cnn").to(device)
    bart_tokenizer = BartTokenizer.from_pretrained("facebook/bart-large-cnn")
    input_text = ' '.join(input_text.split())
    input_tokenized = bart_tokenizer.encode(
        input_text, return_tensors='pt').to(device)
    summary_ids = bart_model.generate(input_tokenized,
                                      num_beams=4, num_return_sequences=1,
                                      no_repeat_ngram_size=2,
                                      length_penalty=1,
                                      min_length=min_len, max_length=mx_len, early_stopping=True)
    output = [bart_tokenizer.decode(
        g, skip_special_tokens=True, clean_up_tokenization_spaces=False) for g in summary_ids]
    return output[0]


nltk.download('punkt')


def my_paraphrase(sentence):
    tokenizer = AutoTokenizer.from_pretrained("Vamsi/T5_Paraphrase_Paws")
    model = AutoModelForSeq2SeqLM.from_pretrained("Vamsi/T5_Paraphrase_Paws")

    sentence = "paraphrase:" + sentence+"</s>"
    encoding = tokenizer.encode_plus(
        sentence, padding=True, return_tensors="pt")
    input_ids, attention_masks = encoding["input_ids"], encoding["attention_mask"]
    outputs = model.generate(input_ids=input_ids, attention_mask=attention_masks, max_length=256,
                             do_sample=True, top_k=120, top_p=0.95, early_stopping=True, num_return_sequences=1)
    output = tokenizer .decode(
        outputs[0], skip_special_tokens=True, clean_up_tokenization_spaces=True)
    return (output)


def get_response(input_text, num_return_sequences):
    model_name = 'tuner007/pegasus_paraphrase'
    torch_device = 'cpu'
    tokenizer = PegasusTokenizer.from_pretrained(model_name)
    model = PegasusForConditionalGeneration.from_pretrained(
        model_name).to(torch_device)

    batch = tokenizer.prepare_seq2seq_batch([input_text], truncation=True, padding='longest', max_length=500,
                                            return_tensors="pt").to(torch_device)
    translated = model.generate(**batch, max_length=500, num_beams=10, num_return_sequences=num_return_sequences,
                                temperature=1.5)
    tgt_text = tokenizer.batch_decode(translated, skip_special_tokens=True)
    return tgt_text


def get_response_from_text(context):
    splitter = SentenceSplitter(language='en')
    sentence_list = splitter.split(context)

    paraphrase = []

    for i in sentence_list:
        a = get_response(i, 1)
        paraphrase.append(a)
    paraphrase2 = [' '.join(x) for x in paraphrase]
    paraphrase3 = [' '.join(x for x in paraphrase2)]
    paraphrased_text = str(paraphrase3).strip('[]').strip("'")
    model = PunctuationModel()

    inputText = paraphrased_text.replace("<n>", "")

    inputText = model.restore_punctuation(inputText)

    my_tool = language_tool_python.LanguageTool('en-US')

    paraphrased_text = my_tool.correct(inputText)
    return paraphrased_text


def genMeta(text):
    kw_extractor = yake.KeywordExtractor()
    language = "en"
    max_ngram_size = 3
    deduplication_threshold = 0.9
    numOfKeywords = 20
    custom_kw_extractor = yake.KeywordExtractor(
        lan=language, n=max_ngram_size, dedupLim=deduplication_threshold, top=numOfKeywords, features=None)
    keywords = custom_kw_extractor.extract_keywords(text)
    retList = []
    for k in keywords:
        retList.append(k[0])
    return retList


def GenArticleHTML(text):
    html = "<html><head></head><body>"+text+"</body></html>"
    soup = BeautifulSoup(html)

    print("BRAT RUNNING")
    desc = Brat(text, 50, 90)
    desc = Brat(text, 50, 90)
    # desc = "text"
    print(desc)

    print("BRAT RUNNING")
    title = Brat(desc, 10, 30)
    title = "TestTitle"
    print(title)
    metaDict = {
        'content': 'text/html',
        'description': desc,
        'robots': 'index, archive',
        'twitter:card': desc,
        'twitter:image:src': 'https://pbs.twimg.com/ext_tw_video_thumb/560070131976392705/pu/img/TcG_ep5t-iqdLV5R.jpg',
        'og:title': title,
        'keywords': ','.join(genMeta(text))

    }
    # metaTags.attrs[''] =
    # metaTags.attrs[''] = ''
    # metaTags.attrs[''] =
    # metaTags.attrs[''] =
    # metaTags.attrs[''] =
    # separator = ', '
    metaTags = soup.new_tag('meta')

    for i, j in metaDict.items():
        metaTags = soup.new_tag('meta')

        metaTags.attrs['name'] = i
        metaTags.attrs['content'] = j

        # tag = soup.(metaTags)

        soup.head.append(metaTags)
    ts = datetime.now()
    ts = str(int(ts.strftime("%Y%m%d%H%M%S")))
    # metaTags.attrs[''] =
    properties = {
        'article:publisher': 'AssignAritblePulisher',
        'article:modified_time': ts,
        'og:image': 'REPLACEURIMAGE',
        'article:published_time': ts,
        'og:url': 'URL_FORUR_ARTICLE',
        'article:tag': ''.join(genMeta(text)),
        'og:site_name': 'RANDOMSITENAM'
    }

    for i, j in properties.items():
        metaTags = soup.new_tag('meta')
        metaTags.attrs['property'] = i

        metaTags.attrs['content'] = j

        # tag = soup.(metaTags)

        soup.head.append(metaTags)
    ret = {}
    ret['text'] = text
    ret['meta'] = soup.prettify()
    return ret

# print(PegasusModel(Clean_Text(URL_extract(c
#     ("https://www.indiatoday.in/science/story/russia-kamchatka-peninsula-volcanoes-awaken-major-eruption-warning-ash-lava-2299836-2022-11-21")))))


# def gen_image_from_text(inputText, userName):
#     ts = datetime.datetime.now()
#     ts = str(int(ts.strftime("%Y%m%d%H%M%S")))
#     # pipe = StableDiffusionPipeline.from_pretrained("CompVis/stable-diffusion-v1-4", torch_dtype=torch.float16)
#     pipe = StableDiffusionPipeline.from_pretrained(
#         "CompVis/stable-diffusion-v1-4")

#     pipe = pipe.to("cuda")

#     prompt = str(inputText)
#     # image here is in [PIL format](https://pillow.readthedocs.io/en/stable/)
#     image = pipe(prompt).images[0]

#     # Now to display an image you can either save it such as:
#     # image.save(f"image.png")
#     # torch.cuda.memory_summary(device=None, abbreviated=False)

#     # Now to display an image you can either save it such as:
#     torch.cuda.memory_summary(device=None, abbreviated=False)
#     userDir = os.path.exists(fileDIR+"/"+userName)
#     test = ""
#     if userDir == False:
#         test = os.path.join(fileDIR, userName)
#         os.mkdir(test)
#     else:
#         test = os.path.join(fileDIR, userName)+"/" + \
#             userName+"_"+ts+".png"

#     image.save(test)
#     # torch.cuda.memory_summary(device=None, abbreviated=False)
#     return test


def gen_image_from_text(inputeText):

    PROMPT = inputeText

    openai.api_key = "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"

    response = openai.Image.create(
        prompt=PROMPT,
        n=1,
        size="512x512",
    )

    return response["data"][0]["url"]


def GO(inputtext, FileName):
    print("BART RUNIING")
    summarizer = pipeline(
        "summarization", model="facebook/bart-large-cnn", truncation=True)
    result = summarizer(inputtext, max_length=400, min_length=300)
    # Print the summary
    print("READIGN PDF")
    FILE = path.join(path.dirname(
        path.realpath(__file__)), FileName)
    # ,strip_text=' .\n')
    tables2 = camelot.read_pdf(FILE, pages="all", flavor="lattice")
    # print(tables2)
    number_of_tables = tables2.n

    temp_df = None
    for i in range(number_of_tables):
        if temp_df is None:
            temp_df = tables2[i].df.replace('\\n', ', ')
        else:
            newTemp = tables2[i].df.replace('\\n', ', ')
            data = pd.DataFrame(newTemp)
            temp_df = temp_df.append(data, ignore_index=True)
            # replace('\n',', ', regex=True)
    # if temp_df is not None:
    #     temp_df.columns = Remove_all_line_breaks(temp_df.iloc[0])
    #     temp_df.rows = Remove_all_line_breaks(temp_df)
    #     temp_df = temp_df[1:]
    # temp_df

    print(temp_df)
    return (result[0]["summary_text"], temp_df.to_html(index=False))
    # readPDF = read_pdf(file,pages = 'all', multiple_tables = True, stream = True)

    # # Transform the result into a string table format
    # table = tabulate(readPDF)

    # # Transform the table into dataframe
    # df = pd.read_fwf(io.StringIO(table))

    # # Save the final result as excel file
    # # df.to_csv()
    # return (result[0]["summary_text"],df.to_csv())


def Remove_all_line_breaks(text):
    """  Remove all line breaks from a string """

    return ([i.replace("\n", ", ") for i in text])


# Twitter Analytics

def twitter_sa(hashtag, tweet_count, userName):
    nltk.download('vader_lexicon')
    analyzer = SentimentIntensityAnalyzer()

    # add your API keys and access tokens here
    consumer_key = 'vYe5wPlYi9eiWxjKAWBbApwFD'
    consumer_secret = 'ksmyYxiwwO0LTuw5dRPxyv3YPiXHpVkgPakgeLTCheAqApmODv'
    access_token = '1534432925273313281-cgqCo7SEKl3a8RZKiFrsMZzcuSMdhz'
    access_token_secret = 'K14g0DkUwMcBhek0LuYMJj2dAVT1uCyXvIHD3g4MdQFKx'

    post = []
    negt = []
    vnegt = []
    neut = []
    pos_count = 0
    neg_count = 0
    very_neg_count = 0
    neutral_count = 0
    total_count = 0

    # authenticate with Twitter API
    auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
    auth.set_access_token(access_token, access_token_secret)
    api = tweepy.API(auth)

    search_term = hashtag
    tweet_amount = tweet_count

    # use the Cursor object to iterate through tweets
    # tweets = tweepy.Cursor(api.search_tweets, q=search_term, tweet_mode='extended').items(tweet_amount)
    tweets = api.search_tweets(
        q=search_term, tweet_mode='extended', count=tweet_amount)

    for tweet in tweets:

        # check if the tweet is a retweet
        if tweet.full_text.startswith('RT @'):
            # if the tweet is a retweet, use the retweeted_status attribute
            full_text = tweet.retweeted_status.full_text
        else:
            full_text = tweet.full_text

        scores = analyzer.polarity_scores(full_text)
        tweet_link = "https://twitter.com/i/web/status/" + str(tweet.id)
        tweet_location = tweet.user.location
        total_count = total_count+1
        if scores['compound'] > 0.00:

            post.append(full_text)
            pos_count = pos_count+1
            post.append(tweet_link)
            if tweet_location:
                post.append(tweet_location)
            else:
                post.append("unavailable")
        elif -0.5 < scores['compound'] < 0.00:

            negt.append(full_text)
            neg_count = neg_count+1
            negt.append(tweet_link)
            if tweet_location:
                negt.append(tweet_location)
            else:
                negt.append("unavailable")
        elif scores['compound'] < -0.5:

            vnegt.append(full_text)
            very_neg_count = very_neg_count+1
            vnegt.append(tweet_link)
            if tweet_location:
                vnegt.append(tweet_location)
            else:
                vnegt.append("unavailable")
        elif scores['compound'] == 0:

            neut.append(full_text)
            neutral_count = neutral_count+1
            neut.append(tweet_link)
            if tweet_location:
                neut.append(tweet_location)
            else:
                neut.append("unavailable")

    # create a new Excel workbook
    workbook = openpyxl.Workbook()

    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # create a new sheet for each sentiment category
    positive_sheet = workbook.create_sheet('Positive_tweets')
    negative_sheet = workbook.create_sheet('Negative_tweets')
    very_negative_sheet = workbook.create_sheet('Very_negative_tweets')
    neutral_sheet = workbook.create_sheet('Neutral_tweets')

    # add headers to each sheet
    positive_sheet.append(['Positive_tweet', 'tweet_link', 'tweet_location'])
    negative_sheet.append(['Negative_tweet', 'tweet_link', 'tweet_location'])
    very_negative_sheet.append(
        ['Very_negative_tweet', 'tweet_link', 'tweet_location'])
    neutral_sheet.append(['Neutral_tweet', 'tweet_link', 'tweet_location'])

    # write data to each sheet
    for sheet, tweets in zip([positive_sheet, negative_sheet, very_negative_sheet, neutral_sheet], [post, negt, vnegt, neut]):
        for tweet, tweet_id, tweet_location in zip(tweets[::3], tweets[1::3], tweets[2::3]):
            sheet.append([tweet, tweet_id, tweet_location])

    # save the workbook
    fileName = "files/"+userName+"/twitter_analysis/"+hashtag + ".xlsx"
    workbook.save(fileName)
    return {
        'fileName': fileName,
        'positve': (pos_count/total_count)*100,
        'nagative': (neg_count/total_count)*100,
        'very_negative': (very_neg_count/total_count)*100,
        'nutral': (neutral_count/total_count)*100,
    }
# SPEECH_DECOR


def speaker_diarization(path, spk_num, language, size, userName, meetingName):
    path = path
    uploadDIr = "files/"+str(userName)+"/"

    num_speakers = int(spk_num)  # @param {type:"integer"}

    language = language  # @param ['any', 'English']

    model_size = size  # @param ['tiny', 'base', 'small', 'medium', 'large']

    model_name = model_size
    if language == 'English' and model_size != 'medium':
        model_name += '.en'

    embedding_model = PretrainedSpeakerEmbedding(
        "speechbrain/spkrec-ecapa-voxceleb", device=torch.device("cpu"))

    if path[-3:] != 'wav':
        subprocess.call(['ffmpeg', '-i', path, 'audio.wav', '-y'])
        path = 'audio.wav'

    model = whisper.load_model(model_size)

    result = model.transcribe(path)
    segments = result["segments"]

    with contextlib.closing(wave.open(path, 'r')) as f:
        frames = f.getnframes()
        rate = f.getframerate()
        duration = frames / float(rate)

    audio = Audio()

    def segment_embedding(segment):
        start = segment["start"]
        # Whisper overshoots the end timestamp in the last segment
        end = min(duration, segment["end"])
        clip = Segment(start, end)
        waveform, sample_rate = audio.crop(path, clip)
        return embedding_model(waveform[None])

    embeddings = np.zeros(shape=(len(segments), 192))
    for i, segment in enumerate(segments):
        embeddings[i] = segment_embedding(segment)

    embeddings = np.nan_to_num(embeddings)

    clustering = AgglomerativeClustering(num_speakers).fit(embeddings)
    labels = clustering.labels_
    for i in range(len(segments)):
        segments[i]["speaker"] = 'SPEAKER ' + str(labels[i] + 1)

    def time(secs):
        return timedelta(seconds=round(secs))

    f = open(uploadDIr+"/"+meetingName+"transcript.txt", "w")
    speakers = []
    distinct_speakers = []
    filePath = []
    for (i, segment) in enumerate(segments):
        if i == 0 or segments[i - 1]["speaker"] != segment["speaker"]:
            f.write("\n" + segment["speaker"] + ' ' +
                    str(time(segment["start"])) + '\n')
            temp1 = time(segment["start"])
            temp2 = time(segment["end"])
            speakers.append(segment["speaker"])
        f.write(segment["text"][1:] + ' ')
        if (speakers.count(segment["speaker"]) == 1):
            distinct_speakers.append(segment["speaker"])
            distinct_speakers.append(segment["start"])
            distinct_speakers.append(segment["end"])

    f.close()

    # Regex pattern to match the speaker and dialogue
    pattern = re.compile(r'SPEAKER (\d+) 0:(\d\d):(\d\d)\n(.*)')

    # Open the file for reading
    with open(uploadDIr+"/"+meetingName+"transcript.txt", "r") as f:
        content = f.read()

    # Dictionary to store the dialogues for each speaker
    dialogues = {}

    # Loop through all the matches in the content
    for match in re.finditer(pattern, content):
        speaker = "SPEAKER " + str(match.group(1))
        # time = "0:" + match.group(2) + ":" + match.group(3)
        dialogue = match.group(4)

        # If the speaker is not already in the dictionary, add it and start a list
        if speaker not in dialogues:
            dialogues[speaker] = []

        # Add the dialogue to the list for this speaker
        dialogues[speaker].append(dialogue)

    # Write the dialogues for each speaker to a separate file
    # for speaker, speaker_dialogues in dialogues.items():
        # with open(speaker + ".txt", "w") as f:
        # f.write("\n".join(speaker_dialogues))

    audio_file = AudioSegment.from_file(path, format="wav")
    speaker_names = []
    val = len(distinct_speakers)//3
    for i in range(val):
        j = 1 + i*3
        start_time_str = str(distinct_speakers[j])
        end_time_str = str(distinct_speakers[j+1])
        start_time = float(start_time_str)
        end_time = float(end_time_str)
        segment = audio_file[start_time*1000:end_time*1000]
        p = pyaudio.PyAudio()
        device_index = p.get_default_output_device_info()['index']
        # stream = p.open(output_device_index=device_index,
        #                 format=pyaudio.paInt16, channels=2, rate=44100, output=True)
        # stream.write(segment.raw_data)
        # stream.stop_stream()
        # stream.close()
        # p.terminate()
        name = ''.join(random.choices(string.ascii_uppercase +
                                      string.digits, k=10))
        speaker_names.append(name)
        export_file = uploadDIr+name + ".mp3"
        filePath.append(export_file)
        segment.export(export_file, format="mp3")

    for i in range(val):
        distinct_speakers[i*3] = speaker_names[i]

    # print(distinct_speakers)
    # print(speaker_names)

    i = 0

    for speaker, speaker_dialogues in dialogues.items():
        with open(uploadDIr+speaker_names[i] + ".txt", "w") as f:
            f.write("\n".join(speaker_dialogues))
        i = i+1

    return {
        'distinct_speakers': distinct_speakers,
        'speaker_names': speaker_names,
        'filepath': filePath,
        'transcript_location': uploadDIr+"/"+meetingName+"transcript.txt"
    }
